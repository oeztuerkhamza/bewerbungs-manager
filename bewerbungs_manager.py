#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bewerbungs-Manager – GUI zum Erstellen von Lebenslauf & Anschreiben
Hamza Öztürk · 10.03.2026
"""

import json
import os
import csv
import re
import imaplib
import smtplib
import email
from email.header import decode_header
from email.utils import parsedate_to_datetime, parseaddr
from email.message import EmailMessage
import html
import mimetypes
import subprocess
import sys
import threading
import tkinter as tk
from datetime import date, timedelta
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

# ─── MODULE IMPORTS ──────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

def _ensure_reportlab():
    try:
        import reportlab  # noqa: F401
        return
    except Exception:
        pass

    try:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'reportlab'])
    except Exception as exc:
        raise RuntimeError(
            'ReportLab not installed and auto-install failed. '
            'Please run: pip install reportlab'
        ) from exc

    import reportlab  # noqa: F401


try:
    import generate_anschreiben as gen_a
    import generate_lebenslauf as gen_l
    import generate_kapak as gen_k
except Exception:
    _ensure_reportlab()
    import generate_anschreiben as gen_a
    import generate_lebenslauf as gen_l
    import generate_kapak as gen_k

import ki_assistent as ki

try:
    from PyPDF2 import PdfReader, PdfWriter, Transformation, PageObject
    PDF_LIBS_LOADED = True
except Exception:
    PdfReader, PdfWriter, Transformation, PageObject = None, None, None, None
    PDF_LIBS_LOADED = False

# ─── CONSTANTS ───────────────────────────────────────────────────────────────
PROFILES_DIR  = os.path.join(SCRIPT_DIR, 'bewerbungen')
OUTPUT_DIR    = SCRIPT_DIR
APPLICATIONS_CSV = os.path.join(SCRIPT_DIR, 'Bewerbungen.csv')
APPLICATIONS_XLSX = os.path.join(SCRIPT_DIR, 'Bewerbungen.xlsx')
IMAP_SETTINGS_FILE = os.path.join(SCRIPT_DIR, '.imap_settings.json')
SMTP_SETTINGS_FILE = os.path.join(SCRIPT_DIR, '.smtp_settings.json')
API_KEY_FILE  = os.path.join(SCRIPT_DIR, '.claude_api_key')
MAIL_PDF_DIR = os.path.join(SCRIPT_DIR, 'mail_pdfs')
ARBEITS_ZEUGNIS_PDF = os.path.join(SCRIPT_DIR, 'Zeugnis', 'bewerbung_software_entwickler_herr_öztürk_arbeitszeugnis.pdf')
BERUFSCHULE_ZEUGNIS_PDF = os.path.join(SCRIPT_DIR, 'Zeugnis', 'bewerbung_software_entwickler_herr_öztürk_berufschule_zeugnis.pdf')
DATA_ANALYST_ZERT_PDF = os.path.join(SCRIPT_DIR, 'Zeugnis', 'bewerbung_software_entwickler_herr_öztürk_data_analyst _zertifikate.pdf')
IHK_ZEUGNIS_PDF = os.path.join(SCRIPT_DIR, 'Zeugnis', 'bewerbung_software_entwickler_herr_öztürk_IHK_zeugnis.pdf')
# ─── PREMIUM COLOR PALETTE ───────────────────────────────────────────────────
NAVY       = '#0D1B2A'
NAVY_MID   = '#1B2838'
NAVY_LIGHT = '#274060'
WHITE      = '#FFFFFF'
BG         = '#F0F2F5'
BG2        = '#E8ECF1'
BG_CARD    = '#FFFFFF'
FG         = '#1E293B'
FG_LIGHT   = '#475569'
GRAY       = '#94A3B8'
ACCENT     = '#C9A84C'      # premium gold
ACCENT_HVR = '#B8963F'
ACCENT2    = '#2563EB'       # action blue
ACCENT2_HVR= '#1D4ED8'
SUCCESS    = '#059669'
CARD_BD    = '#E2E8F0'
DIVIDER    = '#CBD5E1'
FONT       = 'Segoe UI'
FONT_MONO  = 'Cascadia Code'


# ─── HELPER ──────────────────────────────────────────────────────────────────
def today_de():
    """Return today's date as DD.MM.YYYY."""
    d = date.today()
    return f'{d.day:02d}.{d.month:02d}.{d.year}'


def safe_filename(text):
    """Turn arbitrary text into a safe filename component."""
    keep = set('abcdefghijklmnopqrstuvwxyz0123456789_-')
    return ''.join(c if c.lower() in keep else '_' for c in text).strip('_')[:60]


# ─── APPLICATION ─────────────────────────────────────────────────────────────
class BewerbungsApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Bewerbungs-Manager  ·  Premium Suite')
        self.configure(bg=BG)
        self.minsize(1080, 860)
        self.resizable(True, True)

        # Centre on screen
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h = 1140, 900
        self.geometry(f'{w}x{h}+{(sw-w)//2}+{(sh-h)//2}')

        self._build_ui()
        self._load_defaults()

    # ── UI BUILDING ──────────────────────────────────────────────────────────
    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use('clam')

        # ── Premium Button Styles ──
        style.configure('Gold.TButton',
                        background=ACCENT, foreground=NAVY,
                        font=(FONT, 10, 'bold'), padding=(16, 8),
                        borderwidth=0)
        style.map('Gold.TButton',
                  background=[('active', ACCENT_HVR), ('pressed', ACCENT_HVR)],
                  foreground=[('active', WHITE)])

        style.configure('Navy.TButton',
                        background=NAVY, foreground=WHITE,
                        font=(FONT, 10, 'bold'), padding=(16, 8),
                        borderwidth=0)
        style.map('Navy.TButton',
                  background=[('active', NAVY_LIGHT), ('pressed', NAVY_LIGHT)])

        style.configure('Accent.TButton',
                        background=ACCENT2, foreground=WHITE,
                        font=(FONT, 10), padding=(12, 7),
                        borderwidth=0)
        style.map('Accent.TButton',
                  background=[('active', ACCENT2_HVR), ('pressed', ACCENT2_HVR)])

        style.configure('Ghost.TButton',
                        background=BG, foreground=FG,
                        font=(FONT, 10), padding=(12, 7),
                        borderwidth=1, relief='solid')
        style.map('Ghost.TButton',
                  background=[('active', CARD_BD), ('pressed', CARD_BD)])

        # ── Label Styles ──
        style.configure('TLabel', background=BG, foreground=FG,
                        font=(FONT, 10))
        style.configure('Header.TLabel', background=NAVY, foreground=WHITE,
                        font=(FONT, 15, 'bold'), padding=(14, 10))
        style.configure('Section.TLabel', background=BG, foreground=NAVY,
                        font=(FONT, 11, 'bold'))
        style.configure('SectionCard.TLabel', background=WHITE, foreground=NAVY,
                        font=(FONT, 11, 'bold'))
        style.configure('CardLabel.TLabel', background=WHITE, foreground=FG,
                        font=(FONT, 10))
        style.configure('Subtle.TLabel', background=BG, foreground=GRAY,
                        font=(FONT, 9))

        # ── Frame Styles ──
        style.configure('TFrame', background=BG)
        style.configure('Card.TFrame', background=WHITE)
        style.configure('NavyFrame.TFrame', background=NAVY)

        # ── Notebook (Premium Tabs) ──
        style.configure('TNotebook', background=BG, borderwidth=0)
        style.configure('TNotebook.Tab',
                        font=(FONT, 10, 'bold'),
                        padding=(20, 10),
                        background=BG2,
                        foreground=FG_LIGHT)
        style.map('TNotebook.Tab',
                  background=[('selected', WHITE)],
                  foreground=[('selected', NAVY)],
                  expand=[('selected', [0, 0, 0, 2])])

        # ── Entry Style ──
        style.configure('TEntry', font=(FONT, 10), padding=6)

        # ── Separator ──
        style.configure('Gold.TSeparator', background=ACCENT)

        # ═══ HEADER (Gradient Canvas) ═══
        hdr_h = 64
        hdr = tk.Canvas(self, height=hdr_h, bg=NAVY, highlightthickness=0)
        hdr.pack(fill='x')
        # Draw subtle gradient stripe at bottom
        for i in range(6):
            alpha_color = self._blend(NAVY, ACCENT, i / 5)
            hdr.create_rectangle(0, hdr_h - 6 + i, 2000, hdr_h - 5 + i,
                                 fill=alpha_color, outline='')
        # Logo / Title
        hdr.create_text(24, hdr_h // 2 - 2, anchor='w',
                        text='◆  BEWERBUNGS-MANAGER',
                        fill=WHITE, font=(FONT, 16, 'bold'))
        hdr.create_text(320, hdr_h // 2 - 2, anchor='w',
                        text='Premium Suite',
                        fill=ACCENT, font=(FONT, 11))

        # ── Gold accent line ──
        tk.Frame(self, bg=ACCENT, height=3).pack(fill='x')

        # ═══ STATUS BAR ═══  (pack bottom elements FIRST so notebook doesn't eat all space)
        status_frame = tk.Frame(self, bg=NAVY, height=28)
        status_frame.pack(fill='x', side='bottom')
        status_frame.pack_propagate(False)
        self._status_var = tk.StringVar(value='◆  Bereit.')
        tk.Label(status_frame, textvariable=self._status_var,
                 bg=NAVY, fg=ACCENT,
                 font=(FONT, 9), anchor='w', padx=12).pack(
                     side='left', fill='both', expand=True)
        tk.Label(status_frame, text='v2.0  Premium',
                 bg=NAVY, fg=GRAY,
                 font=(FONT, 8), anchor='e', padx=12).pack(side='right')

        # ═══ BOTTOM ACTION BAR ═══
        bar = tk.Frame(self, bg=NAVY_MID, height=66)
        bar.pack(fill='x', side='bottom', pady=(6, 0))
        bar.pack_propagate(False)

        inner = tk.Frame(bar, bg=NAVY_MID)
        inner.pack(expand=True)

        ttk.Button(inner, text='📄  Lebenslauf',
                   style='Gold.TButton',
                   command=self._gen_lebenslauf).pack(
                       side='left', padx=5, pady=12)
        ttk.Button(inner, text='✉  Anschreiben',
                   style='Gold.TButton',
                   command=self._gen_anschreiben).pack(
                       side='left', padx=5, pady=12)
        ttk.Button(inner, text='📑  Beide erstellen',
                   style='Gold.TButton',
                   command=self._gen_both).pack(
                       side='left', padx=5, pady=12)
        ttk.Button(inner, text='📦  Bewerbungs-PDF',
                   style='Navy.TButton',
                   command=self._gen_bewerbung_pdf).pack(
                       side='left', padx=5, pady=12)
        ttk.Button(inner, text='📂  Ordner öffnen',
                   style='Ghost.TButton',
                   command=self._open_folder).pack(
                       side='left', padx=5, pady=12)

        # ═══ NOTEBOOK ═══
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=16, pady=(12, 0))

        self._tab_ki        = self._make_tab(nb, '🤖  KI-Assistent')
        self._tab_stelle    = self._make_tab(nb, '📋  Stelle & Firma')
        self._tab_anschr    = self._make_tab(nb, '✍  Anschreiben')
        self._tab_email     = self._make_tab(nb, '📧  E-Mail')
        self._tab_profile   = self._make_tab(nb, '👤  Profile')

        self._build_ki_tab(self._tab_ki)
        self._build_stelle_tab(self._tab_stelle)
        self._build_anschreiben_tab(self._tab_anschr)
        self._build_email_tab(self._tab_email)
        self._build_profile_tab(self._tab_profile)
        self._nb = nb

    @staticmethod
    def _blend(c1, c2, t):
        """Linearly blend two hex colours; t in [0, 1]."""
        r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
        r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
        r = int(r1 + (r2 - r1) * t)
        g = int(g1 + (g2 - g1) * t)
        b = int(b1 + (b2 - b1) * t)
        return f'#{r:02x}{g:02x}{b:02x}'

    def _make_tab(self, nb, title):
        frame = ttk.Frame(nb, style='TFrame')
        nb.add(frame, text=f'  {title}  ')
        return frame

    # ── CARD CONTAINER HELPER ────────────────────────────────────────────
    def _make_card(self, parent, title=None, padx=16, pady=8):
        """Create a white card container with optional title."""
        outer = tk.Frame(parent, bg=BG)
        outer.pack(fill='x', padx=padx, pady=pady)
        # Card with border
        card = tk.Frame(outer, bg=WHITE, highlightbackground=CARD_BD,
                        highlightthickness=1, padx=16, pady=12)
        card.pack(fill='x')
        if title:
            tk.Label(card, text=title, bg=WHITE, fg=NAVY,
                     font=(FONT, 11, 'bold'), anchor='w').pack(
                         anchor='w', pady=(0, 8))
            tk.Frame(card, bg=DIVIDER, height=1).pack(fill='x', pady=(0, 8))
        return card

    def _make_card_grid(self, parent, title=None, padx=16, pady=8):
        """Create a white card container using grid layout internally."""
        outer = tk.Frame(parent, bg=BG)
        # Card with border
        card = tk.Frame(outer, bg=WHITE, highlightbackground=CARD_BD,
                        highlightthickness=1, padx=16, pady=12)
        card.pack(fill='x')
        if title:
            tk.Label(card, text=title, bg=WHITE, fg=NAVY,
                     font=(FONT, 11, 'bold'), anchor='w').grid(
                         row=0, column=0, columnspan=3, sticky='w', pady=(0, 4))
            tk.Frame(card, bg=DIVIDER, height=1).grid(
                row=1, column=0, columnspan=3, sticky='ew', pady=(0, 8))
        return outer, card

    # ── TAB 0: KI-ASSISTENT ─────────────────────────────────────────────
    def _build_ki_tab(self, parent):
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient='vertical',
                                  command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=BG)
        scroll_frame.bind('<Configure>',
                          lambda e: canvas.configure(
                              scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # ── Description card ──
        desc_card = self._make_card(scroll_frame, padx=16, pady=(12, 4))
        tk.Label(desc_card, text=(
            'Stellenanzeige einfügen (URL oder Text) \u2192 '
            'Claude analysiert die Anforderungen und erstellt '
            'maßgeschneiderte Bewerbungsunterlagen auf Basis '
            'deines Lebenslaufs.'),
            bg=WHITE, fg=FG_LIGHT, font=(FONT, 10),
            wraplength=850, justify='left', anchor='w').pack(anchor='w')

        # ── API Key Card ──
        api_outer, api_card = self._make_card_grid(scroll_frame, '🔑  CLAUDE API KEY', padx=16, pady=4)
        api_outer.pack(fill='x', padx=16, pady=4)

        row = 2
        tk.Label(api_card, text='API Key', bg=WHITE, fg=FG,
                 font=(FONT, 10)).grid(
            row=row, column=0, sticky='e', padx=(0, 8), pady=4)
        self._api_key_var = tk.StringVar(value=self._load_api_key())
        key_entry = ttk.Entry(api_card, textvariable=self._api_key_var,
                              width=60, font=(FONT, 10), show='•')
        key_entry.grid(row=row, column=1, sticky='w', padx=(0, 8), pady=4)
        ttk.Button(api_card, text='Speichern', style='Accent.TButton',
                   command=self._save_api_key).grid(
                       row=row, column=2, sticky='w', padx=4, pady=4)
        row += 1

        self._show_key = tk.BooleanVar(value=False)
        def _toggle_key():
            key_entry.configure(show='' if self._show_key.get() else '•')
        tk.Checkbutton(api_card, text='Key anzeigen', variable=self._show_key,
                       command=_toggle_key, bg=WHITE, fg=FG,
                       font=(FONT, 9), activebackground=WHITE,
                       selectcolor=WHITE).grid(
                           row=row, column=1, sticky='w', padx=0, pady=0)

        # ── Stellenanzeige Card ──
        job_outer, job_card = self._make_card_grid(scroll_frame, '📋  STELLENANZEIGE', padx=16, pady=4)
        job_outer.pack(fill='x', padx=16, pady=4)

        row = 2
        tk.Label(job_card, text='URL', bg=WHITE, fg=FG,
                 font=(FONT, 10)).grid(
            row=row, column=0, sticky='e', padx=(0, 8), pady=4)
        self._job_url_var = tk.StringVar()
        ttk.Entry(job_card, textvariable=self._job_url_var, width=70,
                  font=(FONT, 10)).grid(
                      row=row, column=1, columnspan=2, sticky='w',
                      padx=(0, 12), pady=4)
        row += 1

        tk.Label(job_card, text='oder Stellentext direkt einfügen:',
                 bg=WHITE, fg=FG_LIGHT, font=(FONT, 10)).grid(
                      row=row, column=0, columnspan=3, sticky='w',
                      padx=0, pady=(8, 2))
        row += 1
        self._job_text_widget = tk.Text(
            job_card, height=10, width=95, font=(FONT, 10),
            wrap='word', bg='#F8FAFC', fg=FG, relief='solid',
            borderwidth=1, highlightbackground=CARD_BD,
            highlightthickness=0, padx=8, pady=6)
        self._job_text_widget.grid(
            row=row, column=0, columnspan=3, sticky='we',
            padx=0, pady=(0, 6))

        # ── Extra Instructions Card ──
        extra_outer, extra_card = self._make_card_grid(scroll_frame, '💡  ZUSÄTZLICHE HINWEISE', padx=16, pady=4)
        extra_outer.pack(fill='x', padx=16, pady=4)

        row = 2
        self._extra_instr_widget = tk.Text(
            extra_card, height=3, width=95, font=(FONT, 10),
            wrap='word', bg='#F8FAFC', fg=FG, relief='solid',
            borderwidth=1, highlightbackground=CARD_BD,
            highlightthickness=0, padx=8, pady=6)
        self._extra_instr_widget.insert('1.0',
            'z.B.: Betone Docker-Erfahrung stärker, '
            'erwähne Remote-Bereitschaft...')
        self._extra_instr_widget.grid(
            row=row, column=0, columnspan=3, sticky='we',
            padx=0, pady=(0, 6))

        # ── Generate Buttons Card ──
        btn_card = self._make_card(scroll_frame, padx=16, pady=4)
        btn_inner = tk.Frame(btn_card, bg=WHITE)
        btn_inner.pack()
        ttk.Button(btn_inner,
                   text='🤖  KI-Bewerbung generieren',
                   style='Gold.TButton',
                   command=self._ki_generate).pack(side='left', padx=6)
        ttk.Button(btn_inner,
                   text='🤖 + 📄  Generieren & PDFs erstellen',
                   style='Navy.TButton',
                   command=self._ki_generate_and_pdf).pack(side='left', padx=6)

        # ── Log Card ──
        log_outer, log_card = self._make_card_grid(scroll_frame, '📊  LOG', padx=16, pady=(4, 12))
        log_outer.pack(fill='x', padx=16, pady=(4, 12))

        self._ki_log = tk.Text(
            log_card, height=8, width=95, font=(FONT_MONO, 9),
            wrap='word', bg='#0F172A', fg='#E2E8F0', relief='flat',
            borderwidth=0, padx=10, pady=8, state='disabled',
            insertbackground='#E2E8F0')
        self._ki_log.grid(row=2, column=0, columnspan=3, sticky='we',
                          padx=0, pady=(0, 4))

    def _ki_section(self, parent, title, row):
        lbl = tk.Label(parent, text=title, bg=WHITE, fg=NAVY,
                       font=(FONT, 11, 'bold'))
        lbl.grid(row=row, column=0, columnspan=3, sticky='w',
                 padx=12, pady=(14, 4))
        return row + 1

    def _log(self, msg):
        self._ki_log.configure(state='normal')
        self._ki_log.insert('end', msg + '\n')
        self._ki_log.see('end')
        self._ki_log.configure(state='disabled')
        self.update_idletasks()

    def _clear_log(self):
        self._ki_log.configure(state='normal')
        self._ki_log.delete('1.0', 'end')
        self._ki_log.configure(state='disabled')

    # ── API KEY PERSISTENCE ─────────────────────────────────────────────
    @staticmethod
    def _load_api_key():
        # 1) Umgebungsvariable hat Vorrang (GitHub Secret / .env)
        env_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if env_key:
            return env_key
        # 2) Fallback: lokale Datei (wird nicht ins Repository eingecheckt)
        if os.path.isfile(API_KEY_FILE):
            with open(API_KEY_FILE, 'r', encoding='utf-8') as f:
                return f.read().strip()
        return ''

    def _save_api_key(self):
        key = self._api_key_var.get().strip()
        with open(API_KEY_FILE, 'w', encoding='utf-8') as f:
            f.write(key)
        self._status('✓  API Key gespeichert.')

    # ── KI GENERATION ───────────────────────────────────────────────────
    def _ki_generate(self, then_pdf=False):
        api_key = self._api_key_var.get().strip()
        if not api_key:
            messagebox.showwarning('API Key fehlt',
                                  'Bitte zuerst einen Claude API Key eingeben.')
            return

        url = self._job_url_var.get().strip()
        pasted = self._job_text_widget.get('1.0', 'end-1c').strip()
        extra = self._extra_instr_widget.get('1.0', 'end-1c').strip()
        # Clean default placeholder
        if extra.startswith('z.B.:'):
            extra = ''

        if not url and not pasted:
            messagebox.showwarning('Stellenanzeige fehlt',
                                  'Bitte eine URL oder den Stellentext einfügen.')
            return

        self._clear_log()
        self._status('KI arbeitet...')
        self._log('▶ KI-Bewerbungsassistent gestartet...')

        def _worker():
            try:
                # 1) Get job text
                if url:
                    self._log(f'↓ Lade Stellenanzeige von: {url}')
                    job_text = ki.fetch_job_text(url)
                    self._log(f'✓ {len(job_text)} Zeichen extrahiert.')
                else:
                    job_text = pasted
                    self._log(f'✓ Eingefügter Text ({len(job_text)} Zeichen).')

                # 2) Call Claude
                self._log('↑ Sende an Claude API...')
                cfg = ki.call_claude(api_key, job_text, extra)
                self._log('✓ Claude-Antwort erhalten!')

                # 3) Add datum
                cfg['datum'] = today_de()

                # 4) Populate GUI
                self.after(0, lambda: self._apply_ki_result(cfg, then_pdf))

            except Exception as exc:
                self._log(f'✗ Fehler: {exc}')
                self.after(0, lambda: self._status(f'Fehler: {exc}'))

        threading.Thread(target=_worker, daemon=True).start()

    def _ki_generate_and_pdf(self):
        self._ki_generate(then_pdf=True)

    def _apply_ki_result(self, cfg, then_pdf=False):
        self._set_config(cfg)
        self._log('✓ Alle Felder ausgefüllt.')

        # Fill email fields
        email_betreff = cfg.get('email_betreff', '')
        email_text = cfg.get('email_text', '')
        if email_betreff:
            self._email_betreff_var.set(email_betreff)
        if email_text:
            self._email_text_widget.delete('1.0', 'end')
            self._email_text_widget.insert('1.0', email_text)
            self._log('✓ E-Mail-Text erstellt.')

        # Show warnings
        warnungen = cfg.get('warnungen', [])
        if warnungen:
            self._log('')
            self._log('⚠ ACHTUNG – Bitte manuell prüfen:')
            warn_lines = []
            for w in warnungen:
                self._log(f'  ⚠ {w}')
                warn_lines.append(f'• {w}')
            messagebox.showwarning(
                'KI-Hinweise – Bitte prüfen',
                'Folgende Punkte konnten nicht automatisch '
                'ermittelt werden:\n\n' + '\n'.join(warn_lines) +
                '\n\nBitte im Tab "Stelle & Firma" manuell korrigieren.')

        stelle = cfg.get('stelle', '?')
        firma = cfg.get('firma', '?')
        self._status(f'✓  KI fertig: {stelle} bei {firma}')

        # Switch to Stelle tab so user sees the result
        self._nb.select(self._tab_stelle)

        if then_pdf and not warnungen:
            self._log('📄 Erstelle PDFs...')
            self._gen_both()
            self._log('✓ PDFs erstellt und geöffnet.')
        elif then_pdf and warnungen:
            self._log('⚠ PDFs NICHT erstellt – bitte erst Warnungen beheben, '
                      'dann manuell "Beide erstellen" klicken.')

    # ── TAB 1: STELLE & FIRMA ────────────────────────────────────────────────
    def _build_stelle_tab(self, parent):
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient='vertical',
                                  command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=BG)
        scroll_frame.bind('<Configure>',
                          lambda e: canvas.configure(
                              scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Bind mousewheel
        def _on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        canvas.bind_all('<MouseWheel>', _on_mousewheel)

        self.vars = {}

        # ── Stelle Card ──
        stelle_outer, stelle_card = self._make_card_grid(scroll_frame, '📋  STELLE', padx=16, pady=(12, 4))
        stelle_outer.pack(fill='x', padx=16, pady=(12, 4))
        row = 2
        row = self._field_card(stelle_card, 'stelle', 'Stellenbezeichnung',
                          'Fullstack Entwickler', row)
        row = self._field_card(stelle_card, 'bewerbung_email', 'Bewerbungs-E-Mail',
                  'jobs@example.com', row)
        row = self._field_card(stelle_card, 'betreff', 'Betreff-Zeile',
                          'Bewerbung als Fullstack Entwickler – C# / .NET / Angular',
                          row, width=70)
        row = self._field_card(stelle_card, 'datum', 'Datum', today_de(), row)

        # ── Firma Card ──
        firma_outer, firma_card = self._make_card_grid(scroll_frame, '🏢  FIRMA', padx=16, pady=4)
        firma_outer.pack(fill='x', padx=16, pady=4)
        row = 2
        row = self._field_card(firma_card, 'firma', 'Firmenname', 'Musterfirma GmbH', row)
        row = self._field_card(firma_card, 'ansprechpartner', 'Ansprechpartner',
                          'Frau / Herrn Mustermann', row)
        row = self._field_card(firma_card, 'firma_strasse', 'Straße', 'Musterstraße 1', row)
        row = self._field_card(firma_card, 'firma_plz_ort', 'PLZ + Ort',
                          '79098 Freiburg im Breisgau', row)

        # ── Anrede & Anlagen Card ──
        anrede_outer, anrede_card = self._make_card_grid(scroll_frame, '✍  ANREDE & ANLAGEN', padx=16, pady=(4, 12))
        anrede_outer.pack(fill='x', padx=16, pady=(4, 12))
        row = 2
        row = self._field_card(anrede_card, 'anrede', 'Anrede',
                          'Sehr geehrte Damen und Herren,', row, width=50)
        row = self._field_card(anrede_card, 'anlagen', 'Anlagen',
                          'Anschreiben, Lebenslauf, Arbeitszeugnis,  Zeugnisse, Zertifikate',
                          row, width=60)

    # ── TAB 2: ANSCHREIBEN-TEXT ──────────────────────────────────────────────
    def _build_anschreiben_tab(self, parent):
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient='vertical',
                                  command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=BG)
        scroll_frame.bind('<Configure>',
                          lambda e: canvas.configure(
                              scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Info card
        info_card = self._make_card(scroll_frame, padx=16, pady=(12, 4))
        tk.Label(info_card, text='HTML-Tags wie <b>fett</b> sind erlaubt.',
                 bg=WHITE, fg=GRAY, font=(FONT, 9)).pack(anchor='w')

        defaults = gen_a.DEFAULT_CONFIG
        for i in range(1, 6):
            key = f'absatz_{i}'
            outer, card = self._make_card_grid(scroll_frame, None, padx=16, pady=4)
            outer.pack(fill='x', padx=16, pady=4)
            row = 0
            row = self._textarea_card(card, key, f'Absatz {i}',
                                 defaults.get(key, ''), row)

    # ── TAB 3: E-MAIL ───────────────────────────────────────────────────────
    def _build_email_tab(self, parent):
        c = tk.Frame(parent, bg=BG)
        c.pack(fill='both', expand=True, padx=0, pady=0)

        # ── Email compose card ──
        compose_card = self._make_card(c, '📧  BEWERBUNGS-E-MAIL', padx=16, pady=(12, 4))

        tk.Label(compose_card,
                 text='KI generiert diesen Text automatisch. '
                 'Du kannst ihn bearbeiten und dann kopieren.',
                 bg=WHITE, fg=GRAY, font=(FONT, 9)).pack(
                      anchor='w', pady=(0, 10))

        # Betreff
        bf = tk.Frame(compose_card, bg=WHITE)
        bf.pack(fill='x', pady=(0, 8))
        tk.Label(bf, text='Betreff:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._email_betreff_var = tk.StringVar(
            value='Bewerbung als Fullstack Entwickler')
        ttk.Entry(bf, textvariable=self._email_betreff_var,
                  width=70, font=(FONT, 10)).pack(
                      side='left', fill='x', expand=True)

        # Email body
        self._email_text_widget = tk.Text(
            compose_card, height=14, font=(FONT, 10), wrap='word',
            bg='#F8FAFC', fg=FG, relief='solid', borderwidth=1,
            highlightbackground=CARD_BD, highlightthickness=0,
            padx=10, pady=8)
        self._email_text_widget.insert('1.0',
            'Sehr geehrte Damen und Herren,\n\n'
            'anbei übersende ich Ihnen meine Bewerbungsunterlagen '
            'für die ausgeschriebene Stelle als Fullstack Entwickler.\n\n'
            'Ich freue mich auf Ihre Rückmeldung und stehe für '
            'ein persönliches Gespräch gerne zur Verfügung.\n\n'
            'Mit freundlichen Grüßen\n'
            'Hamza Öztürk\n'
            '+49 155 66859378\n'
            'oeztuerk.hamza@web.de')
        self._email_text_widget.pack(fill='both', expand=True, pady=(0, 10))

        # Buttons
        btn_frame = tk.Frame(compose_card, bg=WHITE)
        btn_frame.pack(fill='x')
        ttk.Button(btn_frame, text='📋  Betreff kopieren',
                   style='Ghost.TButton',
                   command=self._copy_email_betreff).pack(
                       side='left', padx=4)
        ttk.Button(btn_frame, text='📋  E-Mail-Text kopieren',
                   style='Accent.TButton',
                   command=self._copy_email_text).pack(
                       side='left', padx=4)
        ttk.Button(btn_frame, text='📋  Alles kopieren',
                   style='Gold.TButton',
                   command=self._copy_email_all).pack(
                       side='left', padx=4)

        # IMAP sync card
        imap_card = self._make_card(c, '📨  POSTEINGANG SYNCHRONISIEREN (IMAP)', padx=16, pady=(8, 4))

        tk.Label(imap_card,
            text='Liest Bewerbungsantworten aus dem Posteingang und '
                 'aktualisiert Bewerbungen.csv automatisch.',
            bg=WHITE, fg=GRAY, font=(FONT, 9)).pack(
                      anchor='w', pady=(0, 10))

        imap_cfg = self._load_imap_settings()

        row1 = tk.Frame(imap_card, bg=WHITE)
        row1.pack(fill='x', pady=(0, 6))
        tk.Label(row1, text='E-Mail:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._imap_email_var = tk.StringVar(value=imap_cfg.get('email', ''))
        ttk.Entry(row1, textvariable=self._imap_email_var,
                  width=40, font=(FONT, 10)).pack(side='left', padx=(0, 12))

        tk.Label(row1, text='IMAP-Server:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._imap_server_var = tk.StringVar(value=imap_cfg.get('server', 'imap.web.de'))
        ttk.Entry(row1, textvariable=self._imap_server_var,
                  width=24, font=(FONT, 10)).pack(side='left')

        row2 = tk.Frame(imap_card, bg=WHITE)
        row2.pack(fill='x', pady=(0, 8))
        tk.Label(row2, text='App-Passwort:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._imap_password_var = tk.StringVar(value=imap_cfg.get('password', ''))
        ttk.Entry(row2, textvariable=self._imap_password_var, show='•',
                  width=30, font=(FONT, 10)).pack(side='left', padx=(0, 12))

        tk.Label(row2, text='Port:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._imap_port_var = tk.StringVar(value=str(imap_cfg.get('port', 993)))
        ttk.Entry(row2, textvariable=self._imap_port_var,
                  width=8, font=(FONT, 10)).pack(side='left', padx=(0, 12))

        self._imap_only_unseen_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            row2,
            text='Nur ungelesene (UNSEEN) E-Mails durchsuchen',
            variable=self._imap_only_unseen_var,
            bg=WHITE,
            fg=FG,
            font=(FONT, 9),
            activebackground=WHITE,
            selectcolor=WHITE).pack(side='left')

        row3 = tk.Frame(imap_card, bg=WHITE)
        row3.pack(fill='x', pady=(0, 4))
        ttk.Button(row3, text='IMAP-Einstellungen speichern',
                   style='Ghost.TButton',
                   command=self._save_imap_settings).pack(side='left', padx=(0, 6))
        ttk.Button(row3, text='E-Mails abrufen + CSV aktualisieren',
                   style='Accent.TButton',
                   command=self._sync_mail_statuses).pack(side='left', padx=(0, 6))
        ttk.Button(row3, text='Alle E-Mails abrufen + CSV aktualisieren',
                   style='Navy.TButton',
                   command=self._sync_all_mail_statuses).pack(side='left')

        row3b = tk.Frame(imap_card, bg=WHITE)
        row3b.pack(fill='x', pady=(6, 0))
        ttk.Button(row3b, text='Alle E-Mails als PDF herunterladen',
               style='Navy.TButton',
               command=self._download_all_mail_pdfs).pack(side='left')

        # SMTP card
        smtp_card = self._make_card(c, '📤  E-MAIL SENDEN (SMTP)', padx=16, pady=(4, 12))

        smtp_cfg = self._load_smtp_settings()
        row4 = tk.Frame(smtp_card, bg=WHITE)
        row4.pack(fill='x', pady=(0, 6))
        tk.Label(row4, text='SMTP-Server:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._smtp_server_var = tk.StringVar(value=smtp_cfg.get('server', 'smtp.web.de'))
        ttk.Entry(row4, textvariable=self._smtp_server_var,
                  width=24, font=(FONT, 10)).pack(side='left', padx=(0, 12))

        tk.Label(row4, text='SMTP-Port:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._smtp_port_var = tk.StringVar(value=str(smtp_cfg.get('port', 587)))
        ttk.Entry(row4, textvariable=self._smtp_port_var,
                  width=8, font=(FONT, 10)).pack(side='left', padx=(0, 12))

        ttk.Button(row4, text='SMTP-Einstellungen speichern',
                   style='Ghost.TButton',
                   command=self._save_smtp_settings).pack(side='left')

        row5 = tk.Frame(smtp_card, bg=WHITE)
        row5.pack(fill='x', pady=(0, 2))
        tk.Label(row5, text='Empfänger-Adresse:', bg=WHITE, fg=NAVY,
                 font=(FONT, 10, 'bold')).pack(
            side='left', padx=(0, 8))
        self._mail_to_var = self.vars.get('bewerbung_email', tk.StringVar(value=''))
        ttk.Entry(row5, textvariable=self._mail_to_var,
                  width=50, font=(FONT, 10)).pack(side='left', padx=(0, 12))
        ttk.Button(row5, text='Bewerbung per E-Mail senden',
                   style='Gold.TButton',
                   command=self._send_application_email).pack(side='left')

    def _copy_email_betreff(self):
        self.clipboard_clear()
        self.clipboard_append(self._email_betreff_var.get())
        self._status('✓  Betreff in Zwischenablage kopiert.')

    def _copy_email_text(self):
        text = self._email_text_widget.get('1.0', 'end-1c').strip()
        self.clipboard_clear()
        self.clipboard_append(text)
        self._status('✓  E-Mail-Text in Zwischenablage kopiert.')

    def _copy_email_all(self):
        betreff = self._email_betreff_var.get()
        text = self._email_text_widget.get('1.0', 'end-1c').strip()
        full = f'Betreff: {betreff}\n\n{text}'
        self.clipboard_clear()
        self.clipboard_append(full)
        self._status('✓  Betreff + E-Mail-Text in Zwischenablage kopiert.')

    @staticmethod
    def _decode_mime_header(value):
        if not value:
            return ''
        parts = []
        for part, enc in decode_header(value):
            if isinstance(part, bytes):
                parts.append(part.decode(enc or 'utf-8', errors='replace'))
            else:
                parts.append(part)
        return ''.join(parts).strip()

    @staticmethod
    def _extract_mail_text(msg):
        chunks = []
        if msg.is_multipart():
            for part in msg.walk():
                ctype = part.get_content_type()
                disp = str(part.get('Content-Disposition', ''))
                if 'attachment' in disp.lower():
                    continue
                if ctype in ('text/plain', 'text/html'):
                    payload = part.get_payload(decode=True)
                    if payload is None:
                        continue
                    charset = part.get_content_charset() or 'utf-8'
                    text = payload.decode(charset, errors='replace')
                    if ctype == 'text/html':
                        text = re.sub(r'<[^>]+>', ' ', text)
                        text = html.unescape(text)
                    chunks.append(text)
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                charset = msg.get_content_charset() or 'utf-8'
                text = payload.decode(charset, errors='replace')
                if msg.get_content_type() == 'text/html':
                    text = re.sub(r'<[^>]+>', ' ', text)
                    text = html.unescape(text)
                chunks.append(text)

        text = '\n'.join(chunks)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    @staticmethod
    def _norm_text(value):
        value = (value or '').lower()
        value = value.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
        value = re.sub(r'[^a-z0-9 ]+', ' ', value)
        value = re.sub(r'\s+', ' ', value)
        return value.strip()

    def _detect_mail_status(self, subject, body):
        text = self._norm_text(f'{subject} {body}')

        sent_words = [
            'ihre bewerbung wurde an',
            'bewerbung wurde an',
            'application was sent to',
            'your application was sent to'
        ]
        invite_words = [
            'vorstellungsgespraech', 'einladung', 'interview',
            'kennenlernen', 'terminvorschlag', 'gespraech'
        ]
        reject_words = [
            'absage', 'leider', 'nicht beruecksichtigen', 'nicht berucksichtigen',
            'haben uns fuer andere', 'gegen sie entschieden', 'muessen ihnen mitteilen'
        ]
        ack_words = [
            'eingang', 'eingangsbestaetigung', 'eingangsbestaetigung',
            'bewerbung erhalten', 'danke fuer ihre bewerbung',
            'vielen dank fuer ihre bewerbung', 'wir haben ihre bewerbung erhalten'
        ]

        if (any(w in text for w in sent_words) and
                ('gesendet' in text or 'sent' in text)):
            return 'Beworben'
        if any(w in text for w in invite_words):
            return 'Einladung zum Vorstellungsgespräch'
        if any(w in text for w in reject_words):
            return 'Absage'
        if any(w in text for w in ack_words):
            return 'Eingangsbestätigung (In Bearbeitung)'
        return None

    @staticmethod
    def _extract_linkedin_company(subject, body):
        patterns = [
            r'Ihre\s+Bewerbung\s+wurde\s+an\s+(.+?)\s+gesendet',
            r'Your\s+application\s+was\s+sent\s+to\s+(.+?)(?:\.|$)',
        ]
        source = f'{subject} {body}'
        for pattern in patterns:
            m = re.search(pattern, source, flags=re.IGNORECASE)
            if m:
                company = m.group(1).strip(' .,-')
                if company:
                    return company
        return ''

    @staticmethod
    def _extract_linkedin_position(body, company):
        if not body:
            return ''

        # Common LinkedIn snippet: "... gesendet. <Position> <Company> · <Ort> ..."
        if company:
            pattern = rf'gesendet\.\s+(.+?)\s+{re.escape(company)}\s+[·|\\-]'
            m = re.search(pattern, body, flags=re.IGNORECASE)
            if m:
                pos = m.group(1).strip(' .,-')
                if pos:
                    return pos

        m = re.search(r'([A-Za-z0-9ÄÖÜäöüß\-/+ ]+\(m/w/d\))', body)
        if m:
            return m.group(1).strip()

        return ''

    @staticmethod
    def _guess_company_from_sender(sender_name, sender_email):
        sender_name = (sender_name or '').strip()
        if sender_name:
            return sender_name

        domain = (sender_email.split('@')[-1] if '@' in sender_email else sender_email).strip().lower()
        if domain.startswith('mail.'):
            domain = domain[5:]
        base = domain.split('.')[0] if domain else 'Unbekannt'
        return base.replace('-', ' ').replace('_', ' ').strip().title()

    def _company_score(self, company, sender_name, sender_email, subject):
        stop = {
            'gmbh', 'ag', 'kg', 'co', 'mbh', 'se', 'ug', 'der', 'die', 'das',
            'und', 'solutions', 'technology', 'technologies', 'deutschland'
        }
        c_tokens = [t for t in self._norm_text(company).split() if len(t) > 2 and t not in stop]
        if not c_tokens:
            return 0

        hay = self._norm_text(f'{sender_name} {sender_email} {subject}')
        return sum(1 for t in c_tokens if t in hay)

    def _match_csv_row(self, rows, sender_name, sender_email, subject):
        best_idx = -1
        best_score = 0
        for i in range(1, len(rows)):
            row = rows[i]
            if len(row) < 2:
                continue
            score = self._company_score(row[0], sender_name, sender_email, subject)
            if score > best_score:
                best_score = score
                best_idx = i
        return best_idx if best_score > 0 else -1

    @staticmethod
    def _mail_date_to_de(mail_date):
        try:
            dt = parsedate_to_datetime(mail_date)
            return dt.strftime('%d.%m.%Y')
        except Exception:
            return today_de()

    @staticmethod
    def _applications_header():
        return ['Firma / Unternehmen', 'Position / Stelle', 'Status / Ergebnis', 'Datum']

    def _read_applications_rows(self):
        rows = []

        # Prefer Excel as source of truth if available.
        if load_workbook is not None and os.path.isfile(APPLICATIONS_XLSX):
            wb = load_workbook(APPLICATIONS_XLSX)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                row = [str(v).strip() if v is not None else '' for v in r]
                if any(cell for cell in row):
                    rows.append(row)

        if not rows and os.path.isfile(APPLICATIONS_CSV):
            with open(APPLICATIONS_CSV, 'r', encoding='utf-8-sig', newline='') as f:
                rows = list(csv.reader(f))

        if not rows:
            rows = [self._applications_header()]

        return rows

    def _write_applications_rows(self, rows):
        with open(APPLICATIONS_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        if Workbook is None:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Bewerbungen'
        for row in rows:
            ws.append(row)
        try:
            wb.save(APPLICATIONS_XLSX)
        except PermissionError:
            # If the main workbook is open in Excel, save a fallback file
            # so data is not lost and the app can continue.
            fallback = os.path.join(SCRIPT_DIR, 'Bewerbungen_backup.xlsx')
            wb.save(fallback)

    @staticmethod
    def _load_imap_settings():
        if os.path.isfile(IMAP_SETTINGS_FILE):
            with open(IMAP_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {'email': '', 'server': 'imap.web.de', 'port': 993, 'password': ''}

    def _save_imap_settings(self):
        settings = {
            'email': self._imap_email_var.get().strip(),
            'server': self._imap_server_var.get().strip() or 'imap.web.de',
            'port': int((self._imap_port_var.get() or '993').strip() or '993'),
            'password': self._imap_password_var.get()
        }
        with open(IMAP_SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
        self._status('✓  IMAP-Einstellungen gespeichert.')

    @staticmethod
    def _load_smtp_settings():
        if os.path.isfile(SMTP_SETTINGS_FILE):
            with open(SMTP_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {'server': 'smtp.web.de', 'port': 587}

    def _save_smtp_settings(self):
        try:
            port = int((self._smtp_port_var.get() or '587').strip() or '587')
        except ValueError:
            messagebox.showwarning('Ungültiger Port', 'SMTP-Port muss eine Zahl sein (z.B. 587).')
            return

        settings = {
            'server': self._smtp_server_var.get().strip() or 'smtp.web.de',
            'port': port
        }
        with open(SMTP_SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
        self._status('✓  SMTP-Einstellungen gespeichert.')

    @staticmethod
    def _attach_file_to_message(msg, path):
        ctype, _ = mimetypes.guess_type(path)
        if ctype:
            maintype, subtype = ctype.split('/', 1)
        else:
            maintype, subtype = 'application', 'octet-stream'

        with open(path, 'rb') as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype,
                               filename=os.path.basename(path))

    def _send_application_email(self):
        sender = self._imap_email_var.get().strip()
        password = self._imap_password_var.get().strip()
        recipient = self._mail_to_var.get().strip()
        smtp_server = self._smtp_server_var.get().strip() or 'smtp.web.de'
        smtp_port_text = self._smtp_port_var.get().strip() or '587'

        if not sender or not password:
            messagebox.showwarning('Fehlende Angaben',
                                   'Bitte Absender-E-Mail und App-Passwort eingeben.')
            return
        if not recipient:
            messagebox.showwarning('Fehlende Angaben',
                                   'Bitte Bewerbungs-E-Mail-Adresse eingeben.')
            return

        try:
            smtp_port = int(smtp_port_text)
        except ValueError:
            messagebox.showwarning('Ungültiger Port', 'SMTP-Port muss eine Zahl sein (z.B. 587).')
            return

        # Persist credentials/settings so user is not asked again next time.
        self._save_imap_settings()
        self._save_smtp_settings()

        subject = self._email_betreff_var.get().strip() or 'Bewerbung'
        body = self._email_text_widget.get('1.0', 'end-1c').strip()
        if not body:
            messagebox.showwarning('Fehlender Text', 'E-Mail-Text darf nicht leer sein.')
            return

        cfg = self._get_config()
        bewerbung_pdf = self._make_output_path('Bewerbung')

        try:
            if not os.path.isfile(bewerbung_pdf):
                self._build_application_pdf(bewerbung_pdf, cfg)
        except Exception as exc:
            messagebox.showerror('PDF-Fehler', f'PDF konnte nicht erstellt werden: {exc}')
            return

        required_files = [bewerbung_pdf]
        missing = [p for p in required_files if not os.path.isfile(p)]
        if missing:
            messagebox.showerror(
                'Fehlende Datei(en)',
                'Folgende Dateien wurden nicht gefunden:\n\n' + '\n'.join(missing))
            return

        self._status('Bewerbungs-E-Mail wird gesendet...')

        def _worker():
            try:
                msg = EmailMessage()
                msg['From'] = sender
                msg['To'] = recipient
                msg['Subject'] = subject
                msg.set_content(body)

                for path in required_files:
                    self._attach_file_to_message(msg, path)

                with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as smtp:
                    smtp.starttls()
                    smtp.login(sender, password)
                    smtp.send_message(msg)

                self._log_application(cfg)
                self.after(0, lambda: self._status('✓  Bewerbungs-E-Mail gesendet.'))
                self.after(0, lambda: messagebox.showinfo(
                    'Erfolgreich gesendet',
                    f'Bewerbungs-E-Mail gesendet an:\n{recipient}'))
            except Exception as exc:
                self.after(0, lambda: self._status(f'E-Mail-Fehler: {exc}'))
                self.after(0, lambda: messagebox.showerror('E-Mail-Sendefehler', str(exc)))

        threading.Thread(target=_worker, daemon=True).start()

    def _sync_mail_statuses(self):
        self._sync_mail_statuses_core(scan_all=False)

    def _sync_all_mail_statuses(self):
        self._sync_mail_statuses_core(scan_all=True)

    def _download_all_mail_pdfs(self):
        email_addr = self._imap_email_var.get().strip()
        password = self._imap_password_var.get().strip()
        server = self._imap_server_var.get().strip() or 'imap.web.de'
        port_text = self._imap_port_var.get().strip() or '993'

        if not email_addr or not password:
            messagebox.showwarning('Fehlende Angaben',
                                   'Bitte E-Mail-Adresse und App-Passwort eingeben.')
            return

        try:
            port = int(port_text)
        except ValueError:
            messagebox.showwarning('Ungültiger Port', 'Port muss eine Zahl sein (z.B. 993).')
            return

        # Persist credentials/settings so user is not asked again next time.
        self._save_imap_settings()

        self._status('Alle E-Mails werden als PDF heruntergeladen...')

        def _worker():
            try:
                saved, scanned, skipped = self._run_mail_pdf_export(
                    email_addr,
                    password,
                    server,
                    port,
                )
                self.after(0, lambda: self._status(
                    f'✓  PDF-Download abgeschlossen. Gespeichert: {saved}, Durchsucht: {scanned}, Übersprungen: {skipped}'))
                self.after(0, lambda: messagebox.showinfo(
                    'PDF-Download abgeschlossen',
                    f'Gespeicherte E-Mails: {saved}\nDurchsuchte E-Mails: {scanned}\nÜbersprungen/Duplikate: {skipped}\n\nOrdner: {MAIL_PDF_DIR}'))
            except Exception as exc:
                self.after(0, lambda: messagebox.showerror('IMAP-Fehler', str(exc)))
                self.after(0, lambda: self._status(f'IMAP-Fehler: {exc}'))

        threading.Thread(target=_worker, daemon=True).start()

    def _sync_mail_statuses_core(self, scan_all=False):
        email_addr = self._imap_email_var.get().strip()
        password = self._imap_password_var.get().strip()
        server = self._imap_server_var.get().strip() or 'imap.web.de'
        port_text = self._imap_port_var.get().strip() or '993'

        if not email_addr or not password:
            messagebox.showwarning('Fehlende Angaben',
                                   'Bitte E-Mail-Adresse und App-Passwort eingeben.')
            return

        try:
            port = int(port_text)
        except ValueError:
            messagebox.showwarning('Ungültiger Port', 'Port muss eine Zahl sein (z.B. 993).')
            return

        # Persist credentials/settings so user is not asked again next time.
        self._save_imap_settings()

        status_msg = 'Alle E-Mails werden synchronisiert...' if scan_all else 'IMAP-Synchronisierung läuft...'
        self._status(status_msg)

        def _worker():
            try:
                changed, added, scanned = self._run_mail_sync(
                    email_addr,
                    password,
                    server,
                    port,
                    self._imap_only_unseen_var.get(),
                    scan_all,
                )
                self.after(0, lambda: self._status(
                    f'✓  E-Mail-Sync abgeschlossen. Durchsucht: {scanned}, Aktualisiert: {changed}, Neu: {added}'))
                self.after(0, lambda: messagebox.showinfo(
                    'IMAP-Synchronisierung abgeschlossen',
                    f'Durchsuchte E-Mails: {scanned}\nAktualisierte Einträge: {changed}\nNeue Einträge: {added}'))
            except Exception as exc:
                self.after(0, lambda: messagebox.showerror('IMAP-Fehler', str(exc)))
                self.after(0, lambda: self._status(f'IMAP-Fehler: {exc}'))

        threading.Thread(target=_worker, daemon=True).start()

    def _run_mail_sync(self, email_addr, password, server, port, only_unseen, scan_all=False):
        rows = self._read_applications_rows()

        changed = 0
        added = 0
        scanned = 0

        conn = imaplib.IMAP4_SSL(server, port)
        try:
            conn.login(email_addr, password)
            conn.select('INBOX', readonly=True)

            if scan_all:
                criteria = 'ALL'
            elif only_unseen:
                criteria = '(UNSEEN)'
            else:
                since_date = (date.today() - timedelta(days=60)).strftime('%d-%b-%Y')
                criteria = f'(SINCE "{since_date}")'

            status, data = conn.search(None, criteria)
            if status != 'OK':
                raise RuntimeError('IMAP-Suche fehlgeschlagen.')

            msg_ids = data[0].split()
            if not scan_all:
                msg_ids = msg_ids[-80:]

            for msg_id in reversed(msg_ids):
                f_status, f_data = conn.fetch(msg_id, '(RFC822)')
                if f_status != 'OK' or not f_data:
                    continue

                raw = f_data[0][1]
                if not raw:
                    continue

                msg = email.message_from_bytes(raw)
                subject = self._decode_mime_header(msg.get('Subject', ''))
                from_raw = self._decode_mime_header(msg.get('From', ''))
                sender_name, sender_email = parseaddr(from_raw)
                body = self._extract_mail_text(msg)
                status_value = self._detect_mail_status(subject, body)
                li_company = self._extract_linkedin_company(subject, body)
                li_position = self._extract_linkedin_position(body, li_company)

                if not status_value:
                    continue

                scanned += 1
                date_value = self._mail_date_to_de(msg.get('Date', ''))

                row_idx = self._match_csv_row(rows, sender_name, sender_email, subject)
                if row_idx < 0 and li_company:
                    row_idx = self._match_csv_row(rows, li_company, sender_email, f'{subject} {li_company}')
                if row_idx >= 0:
                    row = rows[row_idx]
                    row = (row + [''] * 4)[:4]
                    if row[2] != status_value or row[3] != date_value:
                        row[2] = status_value
                        row[3] = date_value
                        rows[row_idx] = row
                        changed += 1
                else:
                    firma = li_company or self._guess_company_from_sender(sender_name, sender_email)
                    pos = li_position or subject.strip() or '(E-Mail Rückmeldung)'
                    rows.append([firma, pos, status_value, date_value])
                    added += 1

            self._write_applications_rows(rows)
        finally:
            try:
                conn.logout()
            except Exception:
                pass

        return changed, added, scanned

    def _run_mail_pdf_export(self, email_addr, password, server, port):
        _ensure_reportlab()
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib.utils import simpleSplit

        os.makedirs(MAIL_PDF_DIR, exist_ok=True)
        run_dir = os.path.join(MAIL_PDF_DIR, date.today().strftime('%Y%m%d'))
        os.makedirs(run_dir, exist_ok=True)

        pdf_path = os.path.join(run_dir, f'all_mails_{date.today().strftime("%Y%m%d")}.pdf')

        saved = 0
        scanned = 0
        skipped = 0
        seen_ids = set()

        width, height = A4
        margin = 18 * mm
        x = margin
        max_width = width - (36 * mm)
        font_name = 'Helvetica'
        font_size = 10
        leading = 12
        y = height - margin

        c = canvas.Canvas(pdf_path, pagesize=A4)
        c.setFont(font_name, font_size)

        conn = imaplib.IMAP4_SSL(server, port)
        try:
            conn.login(email_addr, password)
            mailboxes = self._imap_list_mailboxes(conn)
            for mailbox in mailboxes:
                status, _ = conn.select(mailbox, readonly=True)
                if status != 'OK':
                    continue

                status, data = conn.search(None, 'ALL')
                if status != 'OK':
                    continue

                msg_ids = data[0].split() if data and data[0] else []
                for msg_id in msg_ids:
                    f_status, f_data = conn.fetch(msg_id, '(RFC822)')
                    if f_status != 'OK' or not f_data:
                        continue

                    raw = f_data[0][1]
                    if not raw:
                        continue

                    msg = email.message_from_bytes(raw)
                    scanned += 1

                    msg_id_val = (msg.get('Message-ID') or '').strip()
                    if msg_id_val and msg_id_val in seen_ids:
                        skipped += 1
                        continue
                    if msg_id_val:
                        seen_ids.add(msg_id_val)

                    subject = self._decode_mime_header(msg.get('Subject', ''))
                    from_raw = self._decode_mime_header(msg.get('From', ''))
                    to_raw = self._decode_mime_header(msg.get('To', ''))
                    cc_raw = self._decode_mime_header(msg.get('Cc', ''))
                    date_raw = msg.get('Date', '')
                    body = self._extract_mail_text(msg)
                    attachments = self._extract_attachment_names(msg)

                    if saved > 0:
                        c.showPage()
                        c.setFont(font_name, font_size)
                        y = height - margin

                    lines = [
                        f'=== MAIL {saved + 1} ===',
                        f'Mailbox: {mailbox}',
                        f'Date: {date_raw}',
                        f'From: {from_raw}',
                        f'To: {to_raw}',
                        f'Cc: {cc_raw}',
                        f'Subject: {subject}',
                        '',
                        '--- Body ---',
                        body or '(Empty body)',
                    ]
                    if attachments:
                        lines.extend(['', '--- Attachments ---'])
                        lines.extend(attachments)

                    y = self._write_mail_pdf_block(
                        c,
                        lines,
                        y,
                        height,
                        x,
                        max_width,
                        margin,
                        font_name,
                        font_size,
                        leading,
                        simpleSplit,
                    )
                    saved += 1
        finally:
            try:
                conn.logout()
            except Exception:
                pass

        c.save()

        return saved, scanned, skipped

    @staticmethod
    def _imap_list_mailboxes(conn):
        status, data = conn.list()
        if status != 'OK' or not data:
            return ['INBOX']

        mailboxes = []
        for raw in data:
            if not raw:
                continue
            line = raw.decode('utf-8', errors='replace')
            flags, name = BewerbungsApp._parse_imap_list_line(line)
            if '\\Noselect' in flags:
                continue
            if name:
                mailboxes.append(name)

        return mailboxes or ['INBOX']

    @staticmethod
    def _parse_imap_list_line(line):
        # Example: (\HasNoChildren) "/" "INBOX"
        m = re.match(r'\((?P<flags>[^)]*)\)\s+"(?P<delim>.*?)"\s+"(?P<name>.*)"', line)
        if not m:
            return set(), ''
        flags = set(f.strip() for f in m.group('flags').split())
        name = m.group('name')
        return flags, name

    @staticmethod
    def _extract_attachment_names(msg):
        names = []
        if not msg.is_multipart():
            return names
        for part in msg.walk():
            disp = str(part.get('Content-Disposition', ''))
            if 'attachment' in disp.lower():
                filename = part.get_filename()
                if filename:
                    names.append(filename)
        return names

    @staticmethod
    def _write_mail_pdf_block(canvas_obj, lines, y, page_height, x, max_width,
                              margin, font_name, font_size, leading, simple_split):
        for line in lines:
            for wrapped in simple_split(line, font_name, font_size, max_width):
                if y < margin:
                    canvas_obj.showPage()
                    canvas_obj.setFont(font_name, font_size)
                    y = page_height - margin
                canvas_obj.drawString(x, y, wrapped)
                y -= leading
        return y

    # ── TAB 4: PROFILE ──────────────────────────────────────────────────────
    def _build_profile_tab(self, parent):
        c = tk.Frame(parent, bg=BG)
        c.pack(fill='both', expand=True, padx=0, pady=0)

        profile_card = self._make_card(c, '👤  Gespeicherte Bewerbungs-Profile', padx=16, pady=12)

        list_frame = tk.Frame(profile_card, bg=WHITE)
        list_frame.pack(fill='both', expand=True)

        self._profile_list = tk.Listbox(
            list_frame, font=(FONT, 10), selectmode='single',
            bg='#F8FAFC', fg=FG, selectbackground=ACCENT,
            selectforeground=NAVY, relief='solid', borderwidth=1,
            highlightbackground=CARD_BD, highlightthickness=0,
            activestyle='none')
        self._profile_list.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(list_frame, orient='vertical',
                           command=self._profile_list.yview)
        sb.pack(side='right', fill='y')
        self._profile_list.configure(yscrollcommand=sb.set)

        btn_frame = tk.Frame(profile_card, bg=WHITE)
        btn_frame.pack(fill='x', pady=(10, 0))
        ttk.Button(btn_frame, text='💾  Profil speichern',
                   style='Gold.TButton',
                   command=self._save_profile).pack(side='left', padx=4)
        ttk.Button(btn_frame, text='📂  Profil laden',
                   style='Accent.TButton',
                   command=self._load_profile).pack(side='left', padx=4)
        ttk.Button(btn_frame, text='🗑  Profil löschen',
                   style='Ghost.TButton',
                   command=self._delete_profile).pack(side='left', padx=4)

        self._refresh_profiles()

    # ── UI HELPERS ───────────────────────────────────────────────────────────
    def _section(self, parent, title, row):
        lbl = tk.Label(parent, text=title, bg=BG, fg=NAVY,
                       font=(FONT, 11, 'bold'))
        lbl.grid(row=row, column=0, columnspan=2, sticky='w',
                 padx=12, pady=(14, 4))
        return row + 1

    def _field(self, parent, key, label, default, row, width=40):
        tk.Label(parent, text=label, bg=BG, fg=FG,
                 font=(FONT, 10)).grid(
            row=row, column=0, sticky='e', padx=(12, 6), pady=3)
        var = tk.StringVar(value=default)
        self.vars[key] = var
        e = ttk.Entry(parent, textvariable=var, width=width,
                      font=(FONT, 10))
        e.grid(row=row, column=1, sticky='w', padx=(0, 12), pady=3)
        return row + 1

    def _field_card(self, parent, key, label, default, row, width=40):
        """Field helper for card (white bg) containers."""
        tk.Label(parent, text=label, bg=WHITE, fg=FG,
                 font=(FONT, 10)).grid(
            row=row, column=0, sticky='e', padx=(0, 8), pady=4)
        var = tk.StringVar(value=default)
        self.vars[key] = var
        e = ttk.Entry(parent, textvariable=var, width=width,
                      font=(FONT, 10))
        e.grid(row=row, column=1, sticky='w', padx=(0, 12), pady=4)
        return row + 1

    def _textarea(self, parent, key, label, default, row, height=5):
        tk.Label(parent, text=label, bg=BG, fg=NAVY,
                 font=(FONT, 11, 'bold')).grid(
            row=row, column=0, columnspan=2, sticky='w',
            padx=12, pady=(10, 2))
        row += 1
        txt = tk.Text(parent, height=height, width=90,
                      font=(FONT, 10), wrap='word',
                      bg='#F8FAFC', fg=FG, relief='solid', borderwidth=1,
                      highlightbackground=CARD_BD, highlightthickness=0,
                      padx=8, pady=6)
        txt.insert('1.0', default)
        txt.grid(row=row, column=0, columnspan=2, sticky='we',
                 padx=12, pady=(0, 6))
        self.vars[key] = txt
        return row + 1

    def _textarea_card(self, parent, key, label, default, row, height=5):
        """Textarea helper for card (white bg) containers."""
        tk.Label(parent, text=label, bg=WHITE, fg=NAVY,
                 font=(FONT, 11, 'bold')).grid(
            row=row, column=0, columnspan=2, sticky='w',
            padx=0, pady=(4, 2))
        row += 1
        txt = tk.Text(parent, height=height, width=90,
                      font=(FONT, 10), wrap='word',
                      bg='#F8FAFC', fg=FG, relief='solid', borderwidth=1,
                      highlightbackground=CARD_BD, highlightthickness=0,
                      padx=8, pady=6)
        txt.insert('1.0', default)
        txt.grid(row=row, column=0, columnspan=2, sticky='we',
                 padx=0, pady=(0, 6))
        self.vars[key] = txt
        return row + 1

    # ── CONFIG GATHERING ─────────────────────────────────────────────────────
    def _get_config(self):
        cfg = {}
        for key, widget in self.vars.items():
            if isinstance(widget, tk.StringVar):
                cfg[key] = widget.get()
            elif isinstance(widget, tk.Text):
                cfg[key] = widget.get('1.0', 'end-1c').strip()
        return cfg

    def _set_config(self, cfg):
        for key, widget in self.vars.items():
            val = cfg.get(key, '')
            if isinstance(widget, tk.StringVar):
                widget.set(val)
            elif isinstance(widget, tk.Text):
                widget.delete('1.0', 'end')
                widget.insert('1.0', val)

    def _load_defaults(self):
        """Populate with merged defaults."""
        defaults = {**gen_a.DEFAULT_CONFIG}
        defaults['datum'] = today_de()
        self._set_config(defaults)

    # ── OUTPUT PATHS ─────────────────────────────────────────────────────────
    def _make_output_path(self, doc_type):
        cfg = self._get_config()
        firma = safe_filename(cfg.get('firma', 'Firma'))
        stelle = safe_filename(cfg.get('stelle', 'Stelle'))
        folder = os.path.join(OUTPUT_DIR, 'bewerbungen', f'{firma} - {stelle}')
        os.makedirs(folder, exist_ok=True)
        if doc_type == 'Anschreiben':
            name = 'bewerbung_software_entwickler_herr_öztürk_anschreiben.pdf'
        elif doc_type == 'Bewerbung':
            name = 'bewerbung_software_entwickler_herr_öztürk.pdf'
        elif doc_type == 'Lebenslauf':
            name = 'bewerbung_software_entwickler_herr_öztürk_lebenslauf.pdf'
        elif doc_type == 'Kapak':
            name = 'bewerbung_software_entwickler_herr_öztürk_deckblatt.pdf'
        else:
            name = f'bewerbung_software_entwickler_herr_öztürk_{doc_type.lower()}.pdf'
        return os.path.join(folder, name)

    # ── GENERATION ───────────────────────────────────────────────────────────
    def _gen_lebenslauf(self):
        cfg = self._get_config()
        out = self._make_output_path('Lebenslauf')
        try:
            gen_l.generate(out, cfg)
            folder = os.path.dirname(out)
            self._status(f'✓  Lebenslauf erstellt: {folder}')
            self._open_pdf(out)
        except Exception as exc:
            messagebox.showerror('Fehler', str(exc))

    def _gen_anschreiben(self):
        cfg = self._get_config()
        out = self._make_output_path('Anschreiben')
        try:
            gen_a.generate(out, cfg)
            self._log_application(cfg)
            folder = os.path.dirname(out)
            self._status(f'✓  Anschreiben erstellt: {folder}')
            self._open_pdf(out)
        except Exception as exc:
            messagebox.showerror('Fehler', str(exc))

    def _gen_both(self):
        cfg = self._get_config()
        try:
            out_l = self._make_output_path('Lebenslauf')
            gen_l.generate(out_l, cfg)
            out_a = self._make_output_path('Anschreiben')
            gen_a.generate(out_a, cfg)
            self._log_application(cfg)
            folder = os.path.dirname(out_l)
            self._status(f'✓  Beide PDFs erstellt in: {os.path.basename(folder)}')
            self._open_pdf(out_l)
            self._open_pdf(out_a)
        except Exception as exc:
            messagebox.showerror('Fehler', str(exc))

    def _gen_bewerbung_pdf(self):
        cfg = self._get_config()
        try:
            out = self._make_output_path('Bewerbung')
            self._build_application_pdf(out, cfg)
            folder = os.path.dirname(out)
            self._status(f'✓  Bewerbungs-PDF erstellt: {os.path.basename(folder)}')
            self._open_pdf(out)
        except Exception as exc:
            messagebox.showerror('Fehler', str(exc))

    @staticmethod
    def _ensure_pdf_merger():
        if PDF_LIBS_LOADED and PdfReader is not None and PdfWriter is not None and Transformation is not None and PageObject is not None:
            return PdfReader, PdfWriter, Transformation, PageObject

        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'PyPDF2'])
        except Exception as exc:
            raise RuntimeError(
                'PyPDF2-Installation fehlgeschlagen. Bitte im Terminal ausführen: '
                'pip install PyPDF2'
            ) from exc

        try:
            from PyPDF2 import PdfReader as _PdfReader, PdfWriter as _PdfWriter, Transformation as _Transformation, PageObject as _PageObject
        except Exception as exc:
            raise RuntimeError('PyPDF2 installiert, aber Import fehlgeschlagen.') from exc

        return _PdfReader, _PdfWriter, _Transformation, _PageObject

    def _build_application_pdf(self, output_path, cfg):
        PdfReader, PdfWriter, Transformation, PageObject = self._ensure_pdf_merger()

        kapak = self._make_output_path('Kapak')
        anschreiben = self._make_output_path('Anschreiben')
        lebenslauf = self._make_output_path('Lebenslauf')

        gen_k.generate(kapak, cfg)
        gen_a.generate(anschreiben, cfg)
        gen_l.generate(lebenslauf, cfg)

        parts = [
            kapak,
            anschreiben,
            lebenslauf,
            ARBEITS_ZEUGNIS_PDF,
            IHK_ZEUGNIS_PDF,
            BERUFSCHULE_ZEUGNIS_PDF,
            DATA_ANALYST_ZERT_PDF,
        ]

        missing = [p for p in parts if not os.path.isfile(p)]
        if missing:
            raise FileNotFoundError('Fehlende Datei(en):\n' + '\n'.join(missing))

        writer = PdfWriter()
        a4_width = 595.276
        a4_height = 841.890

        try:
            for path in parts:
                reader = PdfReader(path)
                for page in reader.pages:
                    orig_w = float(page.mediabox.width)
                    orig_h = float(page.mediabox.height)

                    if orig_w == 0 or orig_h == 0:
                        writer.add_page(page)
                        continue

                    # Zaten A4 boyutundaysa doğrudan ekle
                    if abs(orig_w - a4_width) < 1 and abs(orig_h - a4_height) < 1:
                        writer.add_page(page)
                        continue

                    # A4'e orantılı (en-boy bozmadan) sığdırmak için scale_factor
                    scale = min(a4_width / orig_w, a4_height / orig_h)
                    scaled_w = orig_w * scale
                    scaled_h = orig_h * scale
                    tx = (a4_width - scaled_w) / 2
                    ty = (a4_height - scaled_h) / 2

                    # Boş A4 sayfa oluştur, ölçeklenmiş içeriği üstüne yerleştir
                    blank = PageObject.create_blank_page(width=a4_width, height=a4_height)
                    page.add_transformation(
                        Transformation().scale(scale, scale).translate(tx, ty)
                    )
                    page.mediabox.lower_left = (0, 0)
                    page.mediabox.upper_right = (a4_width, a4_height)
                    blank.merge_page(page)
                    writer.add_page(blank)

            with open(output_path, "wb") as f:
                writer.write(f)
        except Exception as exc:
            raise RuntimeError(f'PDF-Zusammenführungsfehler: {exc}') from exc

    def _log_application(self, cfg):
        """Write or update application status in Bewerbungen table files."""
        firma = (cfg.get('firma') or '').strip()
        stelle = (cfg.get('stelle') or '').strip()
        datum = (cfg.get('datum') or '').strip() or today_de()

        if not firma or not stelle:
            return

        rows = self._read_applications_rows()

        updated = False
        for i in range(1, len(rows)):
            row = rows[i]
            if len(row) < 4:
                row = (row + [''] * 4)[:4]
                rows[i] = row

            if row[0].strip().lower() == firma.lower() and row[1].strip().lower() == stelle.lower():
                row[2] = 'Beworben'
                row[3] = datum
                updated = True
                break

        if not updated:
            rows.append([firma, stelle, 'Beworben', datum])

        self._write_applications_rows(rows)

    # ── PROFILE MANAGEMENT ───────────────────────────────────────────────────
    def _profiles_path(self):
        os.makedirs(PROFILES_DIR, exist_ok=True)
        return PROFILES_DIR

    def _refresh_profiles(self):
        self._profile_list.delete(0, 'end')
        d = self._profiles_path()
        for f in sorted(os.listdir(d)):
            if f.endswith('.json'):
                self._profile_list.insert('end', f[:-5])

    def _save_profile(self):
        cfg = self._get_config()
        firma = cfg.get('firma', 'Firma').strip()
        stelle = cfg.get('stelle', 'Stelle').strip()
        name = f'{firma} – {stelle}'
        fname = safe_filename(name) + '.json'
        path = os.path.join(self._profiles_path(), fname)
        with open(path, 'w', encoding='utf-8') as fp:
            json.dump(cfg, fp, ensure_ascii=False, indent=2)
        self._refresh_profiles()
        self._status(f'✓  Profil gespeichert: {name}')

    def _load_profile(self):
        sel = self._profile_list.curselection()
        if not sel:
            messagebox.showinfo('Hinweis', 'Bitte zuerst ein Profil auswählen.')
            return
        name = self._profile_list.get(sel[0])
        path = os.path.join(self._profiles_path(), name + '.json')
        with open(path, 'r', encoding='utf-8') as fp:
            cfg = json.load(fp)
        self._set_config(cfg)
        self._status(f'✓  Profil geladen: {name}')

    def _delete_profile(self):
        sel = self._profile_list.curselection()
        if not sel:
            messagebox.showinfo('Hinweis', 'Bitte zuerst ein Profil auswählen.')
            return
        name = self._profile_list.get(sel[0])
        path = os.path.join(self._profiles_path(), name + '.json')
        if messagebox.askyesno('Profil löschen',
                               f'"{name}" wirklich löschen?'):
            os.remove(path)
            self._refresh_profiles()
            self._status(f'Profil gelöscht: {name}')

    # ── MISC ─────────────────────────────────────────────────────────────────
    def _status(self, msg):
        self._status_var.set(f'◆  {msg}')
        self.update_idletasks()

    @staticmethod
    def _open_pdf(path):
        if sys.platform == 'win32':
            os.startfile(path)
        elif sys.platform == 'darwin':
            subprocess.Popen(['open', path])
        else:
            subprocess.Popen(['xdg-open', path])

    @staticmethod
    def _open_folder():
        """Open the bewerbungen base folder."""
        folder = os.path.join(OUTPUT_DIR, 'bewerbungen')
        os.makedirs(folder, exist_ok=True)
        if sys.platform == 'win32':
            os.startfile(folder)
        else:
            subprocess.Popen(['xdg-open', folder])


# ─── ENTRY POINT ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    app = BewerbungsApp()
    app.mainloop()
